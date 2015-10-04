import xlwt
import requests
import ast
import datetime
from openpyxl import Workbook

now = datetime.datetime.now()

"""
This takes a list of PLOS article dois and will produce a excel spreadsheet a monthly breakdown of all articles 
The list of articles can be in .txt format separated by linebreaks or csv where the list of articles are in separate cells along the same column.

Requires:
    
    - The excel file creator package xlwt available from https://pypi.python.org/pypi/xlwt
    
    
The function report is provided which successfully provided breakdowns in July 2015
"""

class ALM:
    
    def __init__(self, dictionaryWithSourceNamesAsKeys, listOfSourceNames, doi,title,dateParts):
        self.dic = dictionaryWithSourceNamesAsKeys
        self.sources = listOfSourceNames
        self.doi = str(doi)
        self.title = str(title)
        self.dateParts = dateParts

def getALM(doi,url="http://alm.plos.org/api/events?work_id=http://doi.org/", urlForName = "http://alm.plos.org/api/works/doi:"):
    
    #now add the rest of the url...
    BASE_HEADERS = {'Accept': 'application/json',"version":"6"}
    parameters= {'api_key':'PTAGXPDWSBH_CJRNJE54'}
    urlForName = urlForName + str(doi)
    url += str(doi)
    
        
    if url:
        resp = requests.get(url,
                            params = parameters,
                            headers=BASE_HEADERS)
        if resp.status_code != 200:# check for html errors
            return doi + " returns " + str(resp.status_code) + " error"
                        
        resp = ast.literal_eval(resp.text) ## this built in function maps the text to dictionarys/lists
        
        #we now want to arrange this into a more searchable database, we can go down a level
        
        events = resp["events"] # we only want the events since the rest of the data is only metadata on the request
        
                # events is a list of dictionaries containing all the data for a source
        # the sources names are kept under the "source_id" key in the event's dictionary.. we want this to be searchable by name since the order is not consistent between articles it seems...
        dicByName = {} 
        sources = [] # having a list of the sources so we can iterate easily on it 
        
        #check if events is empty
        if len(events) == 0:
            return doi + " returns empty"
        
        for a in events:
            sources.append(a["source_id"])
            dicByName[a["source_id"]] = a
        
        #do something similar to get the name
        resp = requests.get(urlForName,
                            params = parameters,
                            headers=BASE_HEADERS)
        if resp.status_code != 200:# check for html errors
            return '"' + doi + '"' + " returns " + str(resp.status_code) + " error"
        resp = ast.literal_eval(resp.text)
        name = resp["work"]["title"]
        pubDate = resp["work"]["issued"]["date-parts"][0]
         
        #create an instance of the alm class
        asAlm = ALM(dicByName,sources,doi,name,pubDate)
        
        return asAlm

    else:
        return "invalid doi or url"
    
        
def report(articleFile, output, skipErrorALMs = False, dailyNumbers = False): 
    """
    Takes a file (articleFile) with a list of dois to run. Saves an excel spreadsheet in the containing folder with month by month breakdowns of all the ALMs for that 
    article from all sources available via the PLOS ALM API v6.0
    
    It also will optionally run a report which provides daily data where available.
    
    Articles which return an error will stop the function; however this is bypassed when skipErrorALMs is True
    """
       
    outputPath = output +  "\output.xls"       
    #articleFile = open("C:\Users\Hugh\Desktop\New OpenDocument Spreadsheet.csv")
    #articleList = articleFile.readlines()
    try: 
        articleFile = open(articleFile)
        articleList = articleFile.readlines()
    except:
        articleList = [articleFile]
        
    
    if articleFile == "":
        raise ValueError 
    #cleanup the article list if they have been separated bt commas
    tempList = []
    for a in  articleList:
        a = a.rstrip(",\n")
        tempList.append(a)
    articleList = tempList

    almMasterList = [] #we want a master list to keep all the alms
    
    #collect all the alms!     
    for a in tempList:
        
        tempALM = getALM(a)
        if type(tempALM) == str:
            if skipErrorALMs == False:
                return tempALM
        else:
            almMasterList.append(tempALM)
            print len(almMasterList)
    
    #print almMasterList[0].title
    #print almMasterList[0].dic["pmceuropedata"]["timestamp"]
    #print str((int(almMasterList[0].dic["pmceuropedata"]["timestamp"][0:4]),int(almMasterList[0].dic["pmceuropedata"]["timestamp"][5:7])))
    
    #find the oldest to make the spreadsheet from by counter
    oldestLenght = 0
    oldest = False
    for b in almMasterList:
        #print b.sources
        if len(b.dic["counter"]["by_month"]) > oldestLenght:
            oldestLenght = len(b.dic["counter"]["by_month"])
            oldest = b
            
    
    if oldest == False:
        return("No ALM found to build report")
        
        
    #Now we have the oldest we want to make the xls
    
    book = xlwt.Workbook(encoding="utf-8") 
    
    #set up all the sheets with article titles and dois named by source
    sheetDict = {}
    dateColumnIndexDic = {} # in the form {"sourceName": [(month,year)...]...}
    
    for sor in oldest.sources:
        sheet=book.add_sheet(sor)
        sheet.write(0, 0, "DOI")
        sheet.write(0, 1,"Title")
        
        sheetDict[sor] = sheet
        
        #while we're here lets write all the dates on every page and keep a list of the dates for each source
        months = oldest.dic[sor]["by_month"]
        listOfMonths = []#in the form [(month,year)...]
        columnForMonth = 2
        
        dateIndex = 0
        month = 0
        year = 0
        for date in months:
            if date == months[0]:
                sheet.write(0,columnForMonth, (str(months[dateIndex]["month"]) + "-" +str(months[dateIndex]["year"])))
                listOfMonths.append((months[dateIndex]["month"],months[dateIndex]["year"]))
                month = months[dateIndex]["month"]
                year = months[dateIndex]["year"]
            else:
                if(month!=12):
                    month +=1
                else:
                    month = 1
                    year +=1
                sheet.write(0,columnForMonth, str(month)+"-"+str(year))    
                listOfMonths.append((month,year))
            columnForMonth += 1
            dateIndex +=1
        try:    
            while listOfMonths[-1][1] < int(oldest.dic[sor]["timestamp"][0:4]) or listOfMonths[-1][0] < int(almMasterList[0].dic[sor]["timestamp"][5:7]):#keep adding months if we're not up to date
                #print listOfMonths[-1][1] < int(oldest.dic[sor]["timestamp"][0:4])
                if(month!=12):
                    month +=1
                else:
                    month = 1
                    year +=1
                sheet.write(0,columnForMonth, str(month)+"-"+str(year))    
                listOfMonths.append((month,year))
                columnForMonth += 1
                dateIndex +=1
        except:
            a=a
        #    print sor
        
        dateColumnIndexDic[sor] = listOfMonths
    
    #now lets run through every manuscript and write their data
    workingRow = 1
    
    for met in almMasterList:
        for s in met.sources:
            try:
                workingSheet = sheetDict[s]
            except: #if the source wasn't in the oldest we can just make it
                workingSheet = book.add_sheet(s)
                workingSheet.write(0, 0, "DOI")
                workingSheet.write(0, 1,"Title")
                
                
            workingSheet.write(workingRow,0, met.doi)
            workingSheet.write(workingRow,1, met.title)
            
            #columnInitializer = 2 # reset the columns
            
            columnIndex = False
            #dateList.index(metric.sources['counter'].by_month[0][0]) + columnForMonth
            
            for data in met.dic[s]["by_month"]:
                
                try:
                    columnIndex = dateColumnIndexDic[s].index((data["month"],data["year"])) +2
                    workingSheet.write(workingRow, columnIndex, data["total"])
                except:
                    # need to add missing dates
                    try:#what if it's empty?
                        if len(dateColumnIndexDic.keys()) == 0:
                            columnIndex = 2
                            workingSheet.write(workingRow, columnIndex, data["total"])
                            workingSheet.write(0,columnIndex, str(data["month"])+"-"+str(data["year"]))
                    except:                
                        try:#what if it's missing a more recent date?
                            columnIndex = dateColumnIndexDic[s].index((data["month"]-1,data["year"])) +3 
                            workingSheet.write(workingRow, columnIndex, data["total"])
                            workingSheet.write(0,columnIndex, str(data["month"])+"-"+str(data["year"]))
                        except:
                            return "Unable to write ALM at: " + str(met.doi) + "_" + str(s) + "_" + str(data["month"]-1)+"-"+str(data["year"])
        
            
        workingRow+=1#move to the next row for a new article
            
    
    
        try:
            book.save(outputPath) # the path to save the output
        except:
            return "Can't access output file"
            
    if dailyNumbers == True:
        
        dailyReport(almMasterList, output, oldest)


def dailyReport (almList, outputPath, oldest):
    
    outputPath += "\dailyOutput.xls"
    
    #set up sheet
    dailyBook = Workbook()
    
    sourceDatesDic = {} # keep a dictionary to find the column indexes for sources
    removalDic = {}
    #get the days for counter and do it for all sources
    
    startYear = oldest.dateParts[0]
    startMonth = oldest.dateParts[1]
    startDay = 1
    
    #check if any of the sources are missing from the oldest sources, if so add them for this purpose
    tempSources = oldest.sources
    for t in almList:
        for k in t.sources:
            if k not in tempSources:
                print k
                tempSources.append(k)
    
    for sor in tempSources:
       year = startYear
       month = startMonth
       day = startDay
        
       sheet=dailyBook.create_sheet()
       sheet.title = sor
       values = ["DOI","Title", "Pub Date","Year","Month","Day"]
       cells = ["A4","B4", "C4","C1","c2","c3"]
       removalDic[sor] = 0     
 
       for c in cells:
           workingCell = sheet.cell(c)
           workingCell.value = values[cells.index(c)]
            
    
       workingColumn = 4
       datesDic = {} # keep a dictionary of the date indexes for this source with the key [year, month, day]
    
       while year <= now.year:
           while month <= 12:
               while day <= 31: 
                    
                   toWrite = [year,month,day]
                   rows = [1,2,3]
                    
                   for d in rows:
                       cell = sheet.cell(row=d,column = workingColumn)
                       cell.value = toWrite[rows.index(d)]
                   datesDic[str([year,month,day])] = workingColumn
                   workingColumn +=1
                   if (month == 2 and day == 28) or (day == 30 and (month == 11 or month ==4 or month ==6 or month ==9)):
                        #print str([year,month,day])
                        break
                   day += 1
               month+=1
               day = 1  
           month = 1
           year+=1
       sourceDatesDic[sor] = datesDic
        
    workingRow = 5    
    
    for met in almList:
        for s in met.sources:  
                
            workingSheet = dailyBook.get_sheet_by_name(s)
            
            columns = [1,2,3]
            content = [met.doi, met.title, str(met.dateParts)]
            for w in columns:
                cell = workingSheet.cell(row=workingRow,column=w)
                cell.value = content[columns.index(w)]
                                
            if len(met.dic[s]["by_day"]) != 0:


                
                for data in met.dic[s]["by_day"]:
                    try:
                        
                        columnIndex = sourceDatesDic[s][str([data["year"],data["month"],data["day"]])]
                        cell = workingSheet.cell(row=workingRow, column = columnIndex)
                        if s == "counter":
                            cell.value = data["pdf"]+data["html"]
                        else:
                            cell.value= data["total"]
                            
                    except:
                        
                        print str([data["year"],data["month"],data["day"]])
            else:
             
                removalDic[s] += 1
                           
           
        workingRow+=1#move to the next row for a new article
  
   
    try:
        print outputPath
        #remove sources with no daily data
        
        for q in tempSources:
            
            if removalDic[q] == len(almList):
                dailyBook.remove_sheet(dailyBook.get_sheet_by_name(q))
        
        dailyBook.remove_sheet(dailyBook.get_sheet_by_name("Sheet"))#remove the default sheet
        dailyBook.save(outputPath)  
    except:
        return "Can't access daily output file"
     
    
    



         
                
        



